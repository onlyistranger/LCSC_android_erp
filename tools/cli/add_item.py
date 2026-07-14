#!/usr/bin/env python3
from __future__ import annotations

import argparse
import base64
import csv
from io import BytesIO
import json
import re
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.parse import urlencode

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from PIL import Image as PilImage


LOCATION_CODE_RE = re.compile(r"^[A-Z]\d+$")
LCSC_VERIFY_KEY = "tg09It3*9h"
LCSC_VERIFY_TOKENS = ("_xvasu", "_xvtsc", "_xvpfs", "_xvpts")
SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_OUTPUT_DIR = SCRIPT_DIR / "output"


class Colors:
    def __init__(self, enabled: bool) -> None:
        self.enabled = enabled

    def paint(self, value: Any, code: str) -> str:
        text = str(value)
        if not self.enabled:
            return text
        return f"\033[{code}m{text}\033[0m"

    def bold(self, value: Any) -> str:
        return self.paint(value, "1")

    def green(self, value: Any) -> str:
        return self.paint(value, "32")

    def yellow(self, value: Any) -> str:
        return self.paint(value, "33")

    def blue(self, value: Any) -> str:
        return self.paint(value, "34")

    def magenta(self, value: Any) -> str:
        return self.paint(value, "35")

    def cyan(self, value: Any) -> str:
        return self.paint(value, "36")

    def dim(self, value: Any) -> str:
        return self.paint(value, "2")


@dataclass
class Component:
    id: int
    part_number: str
    name: str | None
    brand: str | None
    package_name: str | None
    category: str | None
    spec_json: str | None
    description: str | None
    source_url: str | None


@dataclass
class Location:
    id: int
    code: str
    display_name: str | None
    color_hex: str | None
    sort_mode: str | None
    remark: str | None
    created_at: int


@dataclass
class InventoryItem:
    id: int
    component_id: int
    location_id: int
    quantity: int
    last_inbound_at: int
    updated_at: int


@dataclass
class FetchedComponent:
    part_number: str
    name: str | None
    brand: str | None
    package_name: str | None
    category: str | None
    spec_json: str | None
    description: str | None
    source_url: str | None
    image_url: str | None


@dataclass
class InboundRequest:
    part_number: str
    count: int
    location: str | None = None


@dataclass
class InboundResult:
    part_number: str
    count: int
    location_code: str
    component_status: str
    inventory_status: str
    quantity_after: int
    fetched: FetchedComponent | None
    component_row: int
    image_status: str | None = None


@dataclass
class InboundCategoryLocationMapping:
    keywords: tuple[str, ...]
    prefix: str


INBOUND_CATEGORY_LOCATION_MAPPINGS = (
    InboundCategoryLocationMapping(("电阻",), "R"),
    InboundCategoryLocationMapping(("电容",), "C"),
    InboundCategoryLocationMapping(("二极管", "LED", "TVS"), "D"),
    InboundCategoryLocationMapping(("电感",), "L"),
    InboundCategoryLocationMapping(("三极管", "晶体管", "MOS"), "Q"),
    InboundCategoryLocationMapping(("晶振", "振荡器"), "Y"),
    InboundCategoryLocationMapping(("保险丝",), "F"),
    InboundCategoryLocationMapping(("连接器", "接插件"), "J"),
    InboundCategoryLocationMapping(("继电器",), "K"),
    InboundCategoryLocationMapping(("开关", "按键"), "S"),
    InboundCategoryLocationMapping(("传感器",), "T"),
    InboundCategoryLocationMapping(
        ("集成电路", "接口芯片", "逻辑芯片", "放大器", "驱动器", "存储器", "处理器", "单片机"),
        "U",
    ),
)

PACKAGE_PARAMETER_KEYS = {
    "封装",
    "封装规格",
    "商品封装",
    "安装类型",
    "Package",
    "Package / Case",
    "Case",
    "Footprint",
}


class WorkbookError(RuntimeError):
    pass


def main() -> int:
    args = parse_args()
    colors = Colors(enabled=not args.no_color and sys.stdout.isatty())
    workbook_path = Path(args.file).expanduser()
    if not workbook_path.exists():
        raise WorkbookError(f"file not found: {workbook_path}")

    requests_to_add = load_inbound_requests(args)
    now_ms = int(time.time() * 1000)

    workbook = load_workbook(workbook_path)
    try:
        sheets = load_inventory_workbook(workbook)
        location_profiles = build_location_category_lookup(sheets)
        print_location_category_profiles(sheets, location_profiles, colors)
        for request in requests_to_add:
            result = process_inbound_request(
                request=request,
                sheets=sheets,
                location_profiles=location_profiles,
                fallback_location=args.fallback_location.upper(),
                no_fetch=args.no_fetch,
                timeout=args.timeout,
                now_ms=now_ms,
            )
            if not args.dry_run:
                result.image_status = insert_component_preview_image(
                    sheet=sheets.components_sheet,
                    headers=sheets.components_headers,
                    row=result.component_row,
                    image_url=result.fetched.image_url if result.fetched else None,
                    timeout=args.timeout,
                )
            print_inbound_result(
                result=result,
                dry_run=args.dry_run,
                colors=colors,
                show_fetch_note=True,
            )
            location_profiles = build_location_category_lookup(sheets)

        output_path = None
        if not args.dry_run:
            output_path = build_output_path(
                source_path=workbook_path,
                label=build_output_label(requests_to_add),
                output_dir=Path(args.output_dir).expanduser(),
                now_ms=now_ms,
            )
            output_path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(output_path)

        if output_path is not None:
            print(f"saved: {colors.green(output_path)}")
        return 0
    finally:
        workbook.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Add an LCSC item quantity to an Android app backup workbook."
    )
    parser.add_argument("--file", required=True, help="Path to lcsc_inventory_backup_*.xlsx")
    parser.add_argument("--csv", help="CSV file for batch add. Columns: id[,count][,location]")
    parser.add_argument("-id", "--part-number", help="LCSC part number, for example C4316662")
    parser.add_argument("-c", "--count", type=int, help="Quantity to add")
    parser.add_argument(
        "-l",
        "--location",
        help="Override target location. Without this, the script mirrors the app's auto-suggestion logic.",
    )
    parser.add_argument(
        "--fallback-location",
        default="A1",
        help="Fallback location used when category inference cannot decide. Default: A1",
    )
    parser.add_argument("--no-fetch", action="store_true", help="Do not query LCSC for missing component details")
    parser.add_argument("--timeout", type=float, default=10.0, help="LCSC request timeout in seconds")
    parser.add_argument("--dry-run", action="store_true", help="Print the result without saving the workbook")
    parser.add_argument("--no-color", action="store_true", help="Disable ANSI color output")
    parser.add_argument(
        "--output-dir",
        default=str(DEFAULT_OUTPUT_DIR),
        help="Directory for the generated workbook. Default: tools/cli/output",
    )
    return parser.parse_args()


def print_inbound_result(
    result: InboundResult,
    dry_run: bool,
    colors: Colors,
    show_fetch_note: bool,
) -> None:
    action = "DRY RUN" if dry_run else "OK"
    colored_action = colors.yellow(action) if dry_run else colors.green(action)
    print(
        f"{colored_action}: {colors.cyan(result.part_number)} "
        f"{colors.green(f'+{result.count}')} -> {colors.blue(result.location_code)}; "
        f"component={colors.magenta(result.component_status)}, "
        f"inventory={colors.magenta(result.inventory_status)}, "
        f"quantity={colors.green(result.quantity_after)}",
        flush=True,
    )
    if show_fetch_note and result.fetched is None and result.component_status == "created":
        print(
            colors.yellow("note: LCSC fetch was unavailable or disabled; created a minimal component row."),
            flush=True,
        )
    if not dry_run and result.fetched is not None:
        print(f"imagePreview={colors.magenta(result.image_status)}", flush=True)
    if (
        not dry_run
        and result.image_status == "missing-url"
        and result.fetched is not None
        and result.fetched.source_url
    ):
        print(
            colors.yellow(
                "note: no image URL found in LCSC data; "
                f"app field productVO.breviaryImageUrl is empty; sourceUrl={result.fetched.source_url}"
            ),
            flush=True,
        )


def load_inbound_requests(args: argparse.Namespace) -> list[InboundRequest]:
    if args.csv:
        if args.part_number:
            raise WorkbookError("--csv cannot be combined with -id/--part-number")
        default_count = validate_quantity(args.count, label="--count") if args.count is not None else 0
        requests_to_add = read_inbound_requests_csv(Path(args.csv).expanduser(), default_count=default_count)
    else:
        if not args.part_number or args.count is None:
            raise WorkbookError("provide either --csv or both -id/--part-number and -c/--count")
        requests_to_add = [
            InboundRequest(
                part_number=normalize_part_number(args.part_number),
                count=validate_quantity(args.count, label=args.part_number),
                location=args.location.strip().upper() if args.location else None,
            )
        ]

    if not requests_to_add:
        raise WorkbookError("no inbound rows to process")
    return requests_to_add


def read_inbound_requests_csv(csv_path: Path, default_count: int) -> list[InboundRequest]:
    if not csv_path.exists():
        raise WorkbookError(f"csv file not found: {csv_path}")
    with csv_path.open("r", encoding="utf-8-sig", newline="") as file:
        rows = list(csv.reader(file, csv.excel))

    rows = [row for row in rows if any(cell.strip() for cell in row)]
    if not rows:
        return []

    first_row = [cell.strip() for cell in rows[0]]
    if looks_like_csv_header(first_row):
        return parse_headered_csv_rows(first_row, rows[1:], default_count)
    return parse_positional_csv_rows(rows, default_count)


def looks_like_csv_header(row: list[str]) -> bool:
    normalized = {normalize_csv_header(cell) for cell in row}
    part_headers = {"id", "partnumber", "part_number", "lcsc", "lcscid"}
    return bool(normalized & part_headers)


def parse_headered_csv_rows(header: list[str], rows: list[list[str]], default_count: int) -> list[InboundRequest]:
    header_indexes = {normalize_csv_header(name): index for index, name in enumerate(header)}
    part_index = first_existing_header(header_indexes, ("id", "partnumber", "part_number", "lcsc", "lcscid"))
    count_index = first_existing_header(header_indexes, ("count", "quantity", "qty", "c"))
    location_index = first_existing_header(header_indexes, ("location", "locationcode", "location_code", "loc"))
    if part_index is None:
        raise WorkbookError("CSV header must contain a part number column")

    requests_to_add = []
    for row_number, row in enumerate(rows, start=2):
        part_number = get_csv_cell(row, part_index)
        count_text = get_csv_cell(row, count_index) if count_index is not None else ""
        if not part_number and not count_text:
            continue
        requests_to_add.append(
            InboundRequest(
                part_number=normalize_part_number(part_number),
                count=parse_csv_quantity(count_text, default_count, label=f"CSV row {row_number}"),
                location=normalize_optional_location(get_csv_cell(row, location_index) if location_index is not None else None),
            )
        )
    return requests_to_add


def parse_positional_csv_rows(rows: list[list[str]], default_count: int) -> list[InboundRequest]:
    requests_to_add = []
    for row_number, row in enumerate(rows, start=1):
        if not row or not row[0].strip():
            continue
        requests_to_add.append(
            InboundRequest(
                part_number=normalize_part_number(row[0]),
                count=parse_csv_quantity(row[1] if len(row) > 1 else "", default_count, label=f"CSV row {row_number}"),
                location=normalize_optional_location(row[2] if len(row) > 2 else None),
            )
        )
    return requests_to_add


def normalize_csv_header(value: str) -> str:
    return re.sub(r"[^a-z0-9_]", "", value.strip().lower())


def first_existing_header(headers: dict[str, int], names: tuple[str, ...]) -> int | None:
    for name in names:
        if name in headers:
            return headers[name]
    return None


def get_csv_cell(row: list[str], index: int) -> str:
    return row[index].strip() if index < len(row) else ""


def validate_quantity(value: Any, label: str) -> int:
    try:
        quantity = int(str(value).strip())
    except ValueError as error:
        raise WorkbookError(f"invalid quantity for {label}: {value}") from error
    if quantity <= 0:
        raise WorkbookError(f"quantity must be greater than 0 for {label}")
    return quantity


def parse_csv_quantity(value: str, default_count: int, label: str) -> int:
    text = value.strip()
    if not text:
        return default_count
    return validate_quantity(text, label=label)


def normalize_optional_location(value: str | None) -> str | None:
    normalized = value.strip().upper() if value else ""
    return normalized or None


def process_inbound_request(
    request: InboundRequest,
    sheets: WorkbookSheets,
    location_profiles: dict[int, tuple[str | None, str | None]],
    fallback_location: str,
    no_fetch: bool,
    timeout: float,
    now_ms: int,
) -> InboundResult:
    component = sheets.components_by_part_number.get(request.part_number)
    fetched = None
    if not no_fetch:
        fetched = fetch_lcsc_component(request.part_number, timeout=timeout)

    if component is None:
        component, component_row = append_component(
            sheets=sheets,
            part_number=request.part_number,
            fetched=fetched,
            now_ms=now_ms,
        )
        component_status = "created"
    else:
        component, component_row = enrich_existing_component(
            sheets=sheets,
            component=component,
            fetched=fetched,
            now_ms=now_ms,
        )
        component_status = "updated" if fetched else "existing"

    location_code = request.location or suggest_inbound_location_code(
        component=component,
        sheets=sheets,
        location_profiles=location_profiles,
        fallback_code=fallback_location,
    )
    location = find_or_create_location(
        sheets=sheets,
        code=location_code,
        now_ms=now_ms,
    )
    inventory_status, quantity_after = add_inventory_quantity(
        sheets=sheets,
        component_id=component.id,
        location_id=location.id,
        quantity_delta=request.count,
        now_ms=now_ms,
    )
    return InboundResult(
        part_number=request.part_number,
        count=request.count,
        location_code=location.code,
        component_status=component_status,
        inventory_status=inventory_status,
        quantity_after=quantity_after,
        fetched=fetched,
        component_row=component_row,
    )


@dataclass
class WorkbookSheets:
    storage_sheet: Any
    components_sheet: Any
    inventory_sheet: Any
    storage_headers: dict[str, int]
    components_headers: dict[str, int]
    inventory_headers: dict[str, int]
    locations: list[Location]
    components: list[Component]
    inventory_items: list[InventoryItem]
    locations_by_code: dict[str, Location]
    locations_by_id: dict[int, Location]
    components_by_part_number: dict[str, Component]
    components_by_id: dict[int, Component]


def load_inventory_workbook(workbook: Any) -> WorkbookSheets:
    required_sheets = ("storage_locations", "components", "inventory_items")
    missing_sheets = [name for name in required_sheets if name not in workbook.sheetnames]
    if missing_sheets:
        raise WorkbookError(f"missing sheet(s): {', '.join(missing_sheets)}")

    storage_sheet = workbook["storage_locations"]
    components_sheet = workbook["components"]
    inventory_sheet = workbook["inventory_items"]
    storage_headers = read_headers(storage_sheet)
    components_headers = read_headers(components_sheet)
    inventory_headers = read_headers(inventory_sheet)

    require_headers(
        "storage_locations",
        storage_headers,
        ("id", "code", "displayName", "colorHex", "sortMode", "remark", "createdAt"),
    )
    require_headers(
        "components",
        components_headers,
        (
            "id",
            "partNumber",
            "name",
            "brand",
            "packageName",
            "category",
            "specJson",
            "description",
            "sourceUrl",
            "updatedAt",
        ),
    )
    require_headers(
        "inventory_items",
        inventory_headers,
        ("id", "componentId", "locationId", "quantity", "lastInboundAt", "updatedAt"),
    )

    locations = read_locations(storage_sheet, storage_headers)
    components = read_components(components_sheet, components_headers)
    inventory_items = read_inventory_items(inventory_sheet, inventory_headers)
    return WorkbookSheets(
        storage_sheet=storage_sheet,
        components_sheet=components_sheet,
        inventory_sheet=inventory_sheet,
        storage_headers=storage_headers,
        components_headers=components_headers,
        inventory_headers=inventory_headers,
        locations=locations,
        components=components,
        inventory_items=inventory_items,
        locations_by_code={location.code.upper(): location for location in locations},
        locations_by_id={location.id: location for location in locations},
        components_by_part_number={component.part_number.upper(): component for component in components},
        components_by_id={component.id: component for component in components},
    )


def read_headers(sheet: Any) -> dict[str, int]:
    headers: dict[str, int] = {}
    for column in range(1, sheet.max_column + 1):
        value = sheet.cell(row=1, column=column).value
        if value is not None and str(value).strip():
            headers[str(value).strip()] = column
    return headers


def require_headers(sheet_name: str, headers: dict[str, int], required: tuple[str, ...]) -> None:
    missing = [header for header in required if header not in headers]
    if missing:
        raise WorkbookError(f"{sheet_name} missing column(s): {', '.join(missing)}")


def read_locations(sheet: Any, headers: dict[str, int]) -> list[Location]:
    locations = []
    for row in iter_data_rows(sheet):
        location_id = read_int(sheet, row, headers["id"])
        code = read_text(sheet, row, headers["code"])
        if not location_id or not code:
            continue
        locations.append(
            Location(
                id=location_id,
                code=code.strip().upper(),
                display_name=read_text(sheet, row, headers["displayName"]),
                color_hex=read_text(sheet, row, headers["colorHex"]),
                sort_mode=read_text(sheet, row, headers["sortMode"]),
                remark=read_text(sheet, row, headers["remark"]),
                created_at=read_int(sheet, row, headers["createdAt"]),
            )
        )
    return locations


def read_components(sheet: Any, headers: dict[str, int]) -> list[Component]:
    components = []
    for row in iter_data_rows(sheet):
        component_id = read_int(sheet, row, headers["id"])
        part_number = read_text(sheet, row, headers["partNumber"])
        if not component_id or not part_number:
            continue
        components.append(
            Component(
                id=component_id,
                part_number=normalize_part_number(part_number),
                name=read_text(sheet, row, headers["name"]),
                brand=read_text(sheet, row, headers["brand"]),
                package_name=read_text(sheet, row, headers["packageName"]),
                category=read_text(sheet, row, headers["category"]),
                spec_json=read_text(sheet, row, headers["specJson"]),
                description=read_text(sheet, row, headers["description"]),
                source_url=read_text(sheet, row, headers["sourceUrl"]),
            )
        )
    return components


def read_inventory_items(sheet: Any, headers: dict[str, int]) -> list[InventoryItem]:
    items = []
    for row in iter_data_rows(sheet):
        item_id = read_int(sheet, row, headers["id"])
        component_id = read_int(sheet, row, headers["componentId"])
        location_id = read_int(sheet, row, headers["locationId"])
        if not item_id or not component_id or not location_id:
            continue
        items.append(
            InventoryItem(
                id=item_id,
                component_id=component_id,
                location_id=location_id,
                quantity=read_int(sheet, row, headers["quantity"]),
                last_inbound_at=read_int(sheet, row, headers["lastInboundAt"]),
                updated_at=read_int(sheet, row, headers["updatedAt"]),
            )
        )
    return items


def iter_data_rows(sheet: Any) -> range:
    return range(2, sheet.max_row + 1)


def read_text(sheet: Any, row: int, column: int) -> str | None:
    value = sheet.cell(row=row, column=column).value
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def read_int(sheet: Any, row: int, column: int) -> int:
    value = sheet.cell(row=row, column=column).value
    if value is None or value == "":
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)
    return int(float(str(value).strip()))


def normalize_part_number(value: str) -> str:
    part_number = value.strip().upper()
    if not re.fullmatch(r"C\d+", part_number):
        raise WorkbookError(f"invalid LCSC part number: {value}")
    return part_number


def append_component(
    sheets: WorkbookSheets,
    part_number: str,
    fetched: FetchedComponent | None,
    now_ms: int,
) -> tuple[Component, int]:
    next_id = next_numeric_id(component.id for component in sheets.components)
    row = sheets.components_sheet.max_row + 1
    component = Component(
        id=next_id,
        part_number=part_number,
        name=fetched.name if fetched else None,
        brand=fetched.brand if fetched else None,
        package_name=fetched.package_name if fetched else None,
        category=fetched.category if fetched else None,
        spec_json=fetched.spec_json if fetched else None,
        description=fetched.description if fetched else None,
        source_url=fetched.source_url if fetched else None,
    )
    write_component_row(sheets.components_sheet, sheets.components_headers, row, component, now_ms)
    sheets.components.append(component)
    sheets.components_by_part_number[part_number] = component
    sheets.components_by_id[next_id] = component
    return component, row


def enrich_existing_component(
    sheets: WorkbookSheets,
    component: Component,
    fetched: FetchedComponent | None,
    now_ms: int,
) -> tuple[Component, int]:
    row = find_row_by_id(sheets.components_sheet, sheets.components_headers["id"], component.id)
    if fetched is None:
        return component, row
    updated = Component(
        id=component.id,
        part_number=component.part_number,
        name=component.name or fetched.name,
        brand=component.brand or fetched.brand,
        package_name=component.package_name or fetched.package_name,
        category=component.category or fetched.category,
        spec_json=component.spec_json or fetched.spec_json,
        description=component.description or fetched.description,
        source_url=component.source_url or fetched.source_url,
    )
    if updated == component:
        return component, row
    write_component_row(sheets.components_sheet, sheets.components_headers, row, updated, now_ms)
    replace_by_id(sheets.components, updated)
    sheets.components_by_part_number[updated.part_number] = updated
    sheets.components_by_id[updated.id] = updated
    return updated, row


def write_component_row(
    sheet: Any,
    headers: dict[str, int],
    row: int,
    component: Component,
    now_ms: int,
) -> None:
    values = {
        "id": component.id,
        "partNumber": component.part_number,
        "name": component.name,
        "brand": component.brand,
        "packageName": component.package_name,
        "category": component.category,
        "specJson": component.spec_json,
        "description": component.description,
        "sourceUrl": component.source_url,
        "updatedAt": now_ms,
    }
    for header, value in values.items():
        sheet.cell(row=row, column=headers[header], value=value)


def insert_component_preview_image(
    sheet: Any,
    headers: dict[str, int],
    row: int,
    image_url: str | None,
    timeout: float,
) -> str:
    image_column = headers.get("imagePreview")
    if image_column is None:
        return "missing-column"
    if not image_url:
        return "missing-url"

    image_bytes = download_image_bytes(image_url, timeout=timeout)
    if image_bytes is None:
        return "download-failed"

    workbook_image_bytes = prepare_workbook_image_bytes(image_bytes)
    if workbook_image_bytes is None:
        return "invalid-image"

    image = OpenpyxlImage(BytesIO(workbook_image_bytes))
    zero_based_column = image_column
    zero_based_row = row - 1
    image.anchor = TwoCellAnchor(
        editAs="twoCell",
        _from=AnchorMarker(col=zero_based_column, row=zero_based_row),
        to=AnchorMarker(col=zero_based_column + 1, row=zero_based_row + 1),
    )
    sheet.add_image(image)
    sheet.row_dimensions[row].height = 72
    column_letter = sheet.cell(row=row, column=image_column + 1).column_letter
    sheet.column_dimensions[column_letter].width = 18
    return "embedded"


def download_image_bytes(image_url: str, timeout: float) -> bytes | None:
    normalized_url = normalize_image_url(image_url)
    if normalized_url is None:
        return None
    try:
        session = requests.Session()
        session.trust_env = False
        response = session.get(
            normalized_url,
            timeout=timeout,
            headers={
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"
                ),
                "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
                "Referer": "https://www.szlcsc.com/",
            },
        )
        response.raise_for_status()
    except requests.RequestException:
        return None
    content_type = response.headers.get("Content-Type", "")
    if "image" not in content_type.lower() and not response.content.startswith((b"\x89PNG", b"\xff\xd8", b"GIF")):
        return None
    return response.content


def normalize_image_url(image_url: str) -> str | None:
    normalized = image_url.strip()
    if not normalized:
        return None
    if normalized.startswith("//"):
        return "https:" + normalized
    if normalized.startswith("/"):
        return "https://item.szlcsc.com" + normalized
    if normalized.startswith(("http://", "https://")):
        return normalized
    return None


def prepare_workbook_image_bytes(image_bytes: bytes) -> bytes | None:
    try:
        with PilImage.open(BytesIO(image_bytes)) as image:
            image_format = (image.format or "").upper()
            if image_format in {"JPEG", "PNG", "GIF"}:
                return image_bytes
            converted = image.convert("RGBA")
            output = BytesIO()
            converted.save(output, format="PNG")
            output.seek(0)
            return output.read()
    except Exception:
        return None


def find_or_create_location(sheets: WorkbookSheets, code: str, now_ms: int) -> Location:
    normalized_code = code.strip().upper()
    if not LOCATION_CODE_RE.fullmatch(normalized_code):
        raise WorkbookError(f"invalid location code: {code}")
    existing = sheets.locations_by_code.get(normalized_code)
    if existing is not None:
        return existing

    next_id = next_numeric_id(location.id for location in sheets.locations)
    row = sheets.storage_sheet.max_row + 1
    location = Location(
        id=next_id,
        code=normalized_code,
        display_name=normalized_code,
        color_hex=None,
        sort_mode="",
        remark=None,
        created_at=now_ms,
    )
    values = {
        "id": location.id,
        "code": location.code,
        "displayName": location.display_name,
        "colorHex": location.color_hex,
        "sortMode": location.sort_mode,
        "remark": location.remark,
        "createdAt": location.created_at,
    }
    for header, value in values.items():
        sheets.storage_sheet.cell(row=row, column=sheets.storage_headers[header], value=value)
    sheets.locations.append(location)
    sheets.locations_by_code[normalized_code] = location
    sheets.locations_by_id[next_id] = location
    return location


def add_inventory_quantity(
    sheets: WorkbookSheets,
    component_id: int,
    location_id: int,
    quantity_delta: int,
    now_ms: int,
) -> tuple[str, int]:
    for item in sheets.inventory_items:
        if item.component_id == component_id and item.location_id == location_id:
            row = find_row_by_id(sheets.inventory_sheet, sheets.inventory_headers["id"], item.id)
            item.quantity += quantity_delta
            item.last_inbound_at = now_ms
            item.updated_at = now_ms
            sheets.inventory_sheet.cell(row=row, column=sheets.inventory_headers["quantity"], value=item.quantity)
            sheets.inventory_sheet.cell(row=row, column=sheets.inventory_headers["lastInboundAt"], value=now_ms)
            sheets.inventory_sheet.cell(row=row, column=sheets.inventory_headers["updatedAt"], value=now_ms)
            return "merged", item.quantity

    next_id = next_numeric_id(item.id for item in sheets.inventory_items)
    row = sheets.inventory_sheet.max_row + 1
    item = InventoryItem(
        id=next_id,
        component_id=component_id,
        location_id=location_id,
        quantity=quantity_delta,
        last_inbound_at=now_ms,
        updated_at=now_ms,
    )
    values = {
        "id": item.id,
        "componentId": item.component_id,
        "locationId": item.location_id,
        "quantity": item.quantity,
        "lastInboundAt": item.last_inbound_at,
        "updatedAt": item.updated_at,
    }
    for header, value in values.items():
        sheets.inventory_sheet.cell(row=row, column=sheets.inventory_headers[header], value=value)
    sheets.inventory_items.append(item)
    return "created", item.quantity


def suggest_inbound_location_code(
    component: Component,
    sheets: WorkbookSheets,
    location_profiles: dict[int, tuple[str | None, str | None]],
    fallback_code: str,
) -> str:
    existing_locations = sorted(
        (
            sheets.locations_by_id[item.location_id]
            for item in sheets.inventory_items
            if item.component_id == component.id and item.location_id in sheets.locations_by_id
        ),
        key=lambda location: location.code,
    )
    if existing_locations:
        return existing_locations[0].code

    sorted_locations = sorted(
        sheets.locations,
        key=lambda location: (
            inbound_location_row_index(location.code),
            inbound_location_column_index(location.code),
            location.code,
        ),
    )
    component_category = normalized_inbound_profile_value(component.category)
    if component_category is not None:
        category_matched = [
            location
            for location in sorted_locations
            if location_profiles.get(location.id, (None, None))[0] == component_category
        ]
        if len(category_matched) == 1:
            return category_matched[0].code
        if len(category_matched) > 1:
            component_package_name = normalized_inbound_profile_value(component.package_name)
            if component_package_name is not None:
                for location in category_matched:
                    if location_profiles.get(location.id, (None, None))[1] == component_package_name:
                        return location.code
            return category_matched[0].code

    normalized_category = (component.category or "").strip()
    if not normalized_category:
        return fallback_code

    for mapping in INBOUND_CATEGORY_LOCATION_MAPPINGS:
        if any(keyword.lower() in normalized_category.lower() for keyword in mapping.keywords):
            for location in sorted_locations:
                if location.code.strip().upper().startswith(mapping.prefix):
                    return location.code
            return f"{mapping.prefix}1"

    return fallback_code


def print_location_category_profiles(
    sheets: WorkbookSheets,
    location_profiles: dict[int, tuple[str | None, str | None]],
    colors: Colors,
) -> None:
    print(colors.bold("Location profiles:"))
    sorted_locations = sorted(
        sheets.locations,
        key=lambda location: (
            inbound_location_row_index(location.code),
            inbound_location_column_index(location.code),
            location.code,
        ),
    )
    for location in sorted_locations:
        category, package_name = location_profiles.get(location.id, (None, None))
        category_text = colors.magenta(category) if category else colors.dim("-")
        package_text = colors.cyan(package_name) if package_name else colors.dim("-")
        print(
            f"  {colors.blue(location.code)}: "
            f"category={category_text}, package={package_text}"
        )


def build_location_category_lookup(sheets: WorkbookSheets) -> dict[int, tuple[str | None, str | None]]:
    profiles_by_location: dict[int, list[tuple[str | None, str | None]]] = {}
    for item in sheets.inventory_items:
        component = sheets.components_by_id.get(item.component_id)
        if component is None:
            continue
        profiles_by_location.setdefault(item.location_id, []).append(
            (component.category, component.package_name)
        )

    lookup: dict[int, tuple[str | None, str | None]] = {}
    for location_id, profiles in profiles_by_location.items():
        category = dominant_profile_value(category for category, _ in profiles)
        package_name = dominant_profile_value(package_name for _, package_name in profiles)
        if category is not None or package_name is not None:
            lookup[location_id] = (category, package_name)
    return lookup


def dominant_profile_value(values: Any) -> str | None:
    counts: dict[str, int] = {}
    for value in values:
        normalized = normalized_inbound_profile_value(value)
        if normalized is not None:
            counts[normalized] = counts.get(normalized, 0) + 1
    if not counts:
        return None
    return sorted(counts.items(), key=lambda item: (-item[1], item[0]))[0][0]


def normalized_inbound_profile_value(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = value.strip()
    return normalized.upper() if normalized else None


def inbound_location_row_index(code: str) -> int:
    return ord(code[0].upper()) if code else sys.maxsize


def inbound_location_column_index(code: str) -> int:
    suffix = re.sub(r"^[A-Za-z]+", "", code)
    return int(suffix) if suffix.isdigit() else sys.maxsize


def next_numeric_id(values: Any) -> int:
    max_value = 0
    for value in values:
        max_value = max(max_value, int(value or 0))
    return max_value + 1


def find_row_by_id(sheet: Any, id_column: int, wanted_id: int) -> int:
    for row in iter_data_rows(sheet):
        if read_int(sheet, row, id_column) == wanted_id:
            return row
    raise WorkbookError(f"row id not found: {wanted_id}")


def replace_by_id(items: list[Component], updated: Component) -> None:
    for index, item in enumerate(items):
        if item.id == updated.id:
            items[index] = updated
            return
    items.append(updated)


def build_output_path(source_path: Path, label: str, output_dir: Path, now_ms: int) -> Path:
    timestamp = time.strftime("%Y%m%d_%H%M%S", time.localtime(now_ms / 1000))
    base_name = f"{source_path.stem}_{label}_{timestamp}{source_path.suffix}"
    candidate = output_dir / base_name
    index = 2
    while candidate.exists():
        candidate = output_dir / f"{source_path.stem}_{label}_{timestamp}_{index}{source_path.suffix}"
        index += 1
    return candidate


def build_output_label(requests_to_add: list[InboundRequest]) -> str:
    if len(requests_to_add) == 1:
        return requests_to_add[0].part_number
    return f"batch_{len(requests_to_add)}items"


def fetch_lcsc_component(part_number: str, timeout: float) -> FetchedComponent | None:
    session = requests.Session()
    session.trust_env = False
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"
            ),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
            "Cache-Control": "no-cache",
            "Pragma": "no-cache",
            "Referer": "https://so.szlcsc.com/",
        }
    )
    url = "https://so.szlcsc.com/global.html?" + urlencode({"k": part_number})
    root = fetch_next_data(session, url, timeout)
    if root is None:
        return None
    product_list = (
        root.get("props", {})
        .get("pageProps", {})
        .get("soData", {})
        .get("searchResult", {})
        .get("productRecordList", [])
    )
    normalized = normalize_part_number(part_number)
    for record in product_list:
        product = record.get("productVO") or {}
        product_code = str(product.get("productCode") or "").strip().upper()
        if product_code == normalized:
            return build_fetched_component(record, product, session=session, timeout=timeout)
    return None


def fetch_next_data(session: requests.Session, url: str, timeout: float) -> dict[str, Any] | None:
    try:
        response = session.get(url, timeout=timeout)
        if response.status_code == 203 and looks_like_verification_page(response.text):
            cookie = build_verification_cookie(response.text)
            if cookie:
                name, value = cookie.split("=", 1)
                session.cookies.set(name, value)
                response = session.get(url, timeout=timeout)
        response.raise_for_status()
    except requests.RequestException:
        return None

    soup = BeautifulSoup(response.text, "html.parser")
    script = soup.select_one("script#__NEXT_DATA__")
    if script is None or not script.string:
        return None
    try:
        return json.loads(script.string)
    except json.JSONDecodeError:
        return None


def looks_like_verification_page(html: str) -> bool:
    return any(token in html for token in LCSC_VERIFY_TOKENS)


def build_verification_cookie(html: str) -> str | None:
    xvasu = extract_javascript_variable(html, "_xvasu")
    xvpts = extract_javascript_variable(html, "_xvpts")
    xvpfs = extract_javascript_variable(html, "_xvpfs")
    if not xvasu or not xvpts or not xvpfs:
        return None
    cookie_name = f"{xvpfs}{xvasu}"
    encrypted = rc4(LCSC_VERIFY_KEY.encode(), f"{xvpts}:{xvasu}".encode())
    cookie_value = base64.b64encode(encrypted).decode()
    return f"{cookie_name}={cookie_value}"


def extract_javascript_variable(html: str, name: str) -> str | None:
    match = re.search(rf"var\s+{re.escape(name)}\s*=\s*(.+?);", html)
    if not match:
        return None
    return match.group(1).strip().strip('"').strip("'")


def rc4(key: bytes, value: bytes) -> bytes:
    state = list(range(256))
    index_b = 0
    for index_a in range(256):
        index_b = (index_b + state[index_a] + key[index_a % len(key)]) % 256
        state[index_a], state[index_b] = state[index_b], state[index_a]

    index_a = 0
    index_b = 0
    output = bytearray()
    for byte in value:
        index_a = (index_a + 1) % 256
        index_b = (index_b + state[index_a]) % 256
        state[index_a], state[index_b] = state[index_b], state[index_a]
        key_stream = state[(state[index_a] + state[index_b]) % 256]
        output.append(byte ^ key_stream)
    return bytes(output)


def build_fetched_component(
    record: dict[str, Any],
    product: dict[str, Any],
    session: requests.Session | None = None,
    timeout: float = 10.0,
) -> FetchedComponent:
    search_params = parse_search_params(record.get("paramLinkedMap") or {})
    product_id = text_or_none(product.get("productId"))
    source_url = f"https://item.szlcsc.com/{product_id}.html" if product_id else None
    category = sanitize_search_text(record.get("lightCatalogName")) or sanitize_search_text(product.get("productType"))

    light_product_model = sanitize_search_text(record.get("lightProductModel"))
    product_model = sanitize_search_text(product.get("productModel"))
    light_product_name = sanitize_search_text(record.get("lightProductName"))
    product_name = sanitize_search_text(product.get("productName"))
    name = light_product_model or product_model or normalize_display_name(
        raw_name=light_product_name or product_name,
        fallback_name=category,
        extracted_specs=list(search_params.values()),
    )

    package_name = (
        sanitize_search_text(record.get("lightStandard"))
        or sanitize_search_text(product.get("encapsulationModel"))
        or extract_package_name_from_search_params(search_params)
    )
    spec_json = json.dumps(search_params, ensure_ascii=False, separators=(",", ":")) if search_params else None
    return FetchedComponent(
        part_number=normalize_part_number(str(product.get("productCode") or "")),
        name=name,
        brand=sanitize_search_text(record.get("lightBrandName"))
        or sanitize_search_text(product.get("productGradePlateName")),
        package_name=package_name,
        category=category,
        spec_json=spec_json,
        description=sanitize_search_text(product.get("remark"))
        or sanitize_search_text(record.get("lightProductIntro")),
        source_url=source_url,
        image_url=text_or_none(product.get("breviaryImageUrl")),
    )


def parse_search_params(raw: dict[str, Any]) -> dict[str, str]:
    params: dict[str, str] = {}
    for key, value in raw.items():
        normalized_key = sanitize_search_text(key)
        normalized_value = sanitize_search_text(value)
        if normalized_key and normalized_value:
            params[normalized_key] = normalized_value
    return params


def normalize_display_name(
    raw_name: str | None,
    fallback_name: str | None,
    extracted_specs: list[str],
) -> str | None:
    if not raw_name:
        return fallback_name
    candidate = raw_name
    for spec_value in sorted((value for value in extracted_specs if value), key=len, reverse=True):
        candidate = candidate.replace(spec_value, " ")
    candidate = re.sub(r"\s+", " ", candidate).strip()
    return candidate or fallback_name or raw_name


def extract_package_name_from_search_params(search_params: dict[str, str]) -> str | None:
    for key, value in search_params.items():
        if key.strip() in PACKAGE_PARAMETER_KEYS:
            return value
    for key, value in search_params.items():
        normalized_key = key.strip().lower()
        if any(candidate.lower() in normalized_key for candidate in PACKAGE_PARAMETER_KEYS):
            return value
    return None


def sanitize_search_text(value: Any) -> str | None:
    text = text_or_none(value)
    if text is None:
        return None
    text = BeautifulSoup(text, "html.parser").get_text()
    text = text.replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text or None


def text_or_none(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text == "null":
        return None
    return text


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except WorkbookError as error:
        print(f"error: {error}", file=sys.stderr)
        raise SystemExit(2)

#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import qrcode
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


REPO_ROOT = Path(__file__).resolve().parents[1]
SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT_PATH = REPO_ROOT / "backup" / "lcsc_inventory_backup0423.xlsx"
DEFAULT_OUTPUT_DIR = SCRIPT_DIR / "output"
PREFERRED_SPEC_KEYS = ("电阻类型", "阻值", "容值", "精度", "功率")
BACKGROUND_COLOR = "#F5F7FA"
CARD_COLOR = "#FFFFFF"
BORDER_COLOR = "#DCE3EA"
MEDIA_BG_COLOR = "#EEF2F6"
TITLE_COLOR = "#111827"
SUBTITLE_COLOR = "#5B6470"
BODY_COLOR = "#111827"


@dataclass(frozen=True)
class ComponentRecord:
    component_id: str
    part_number: str
    mpn: str | None
    name: str | None
    brand: str | None
    package_name: str | None
    category: str | None
    description: str | None
    source_url: str | None
    specifications: dict[str, str]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate LCSC material QR label images from backup workbook components sheet."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT_PATH,
        help="Workbook path. Default: ./backup/lcsc_inventory_backup0423.xlsx",
    )
    parser.add_argument(
        "--sheet",
        default="components",
        help="Sheet name to read. Default: components",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="Directory for generated PNG files. Default: tools/output",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Only process the first N valid component rows.",
    )
    parser.add_argument(
        "--part-number",
        default=None,
        help="Only process one specific part number.",
    )
    return parser.parse_args()


def normalize_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_part_number(value: object) -> str | None:
    text = normalize_text(value)
    return text.upper() if text else None


def parse_specifications(raw_value: object) -> dict[str, str]:
    text = normalize_text(raw_value)
    if not text:
        return {}
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        return {}
    if not isinstance(parsed, dict):
        return {}
    return {
        str(key).strip(): str(value).strip()
        for key, value in parsed.items()
        if str(key).strip() and str(value).strip() and str(value).strip() != "null"
    }


def resolve_input_path(user_input: Path) -> Path:
    candidate = user_input if user_input.is_absolute() else REPO_ROOT / user_input
    if candidate.exists():
        return candidate.resolve()
    raise FileNotFoundError(f"Workbook not found: {candidate}")


def resolve_output_path(user_output: Path) -> Path:
    return user_output if user_output.is_absolute() else (SCRIPT_DIR / user_output).resolve()


def load_components(workbook_path: Path, sheet_name: str) -> list[ComponentRecord]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            available = ", ".join(workbook.sheetnames)
            raise KeyError(f"Sheet '{sheet_name}' not found. Available: {available}")
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return []
        header_names = [normalize_text(header) or "" for header in headers]
        components: list[ComponentRecord] = []
        for row in rows:
            row_map = {
                header_names[index]: row[index] if index < len(row) else None
                for index in range(len(header_names))
            }
            part_number = normalize_part_number(row_map.get("partNumber"))
            if not part_number:
                continue
            components.append(
                ComponentRecord(
                    component_id=str(row_map.get("id")).strip(),
                    part_number=part_number,
                    mpn="",
                    name=normalize_text(row_map.get("name")),
                    brand=normalize_text(row_map.get("brand")),
                    package_name=normalize_text(row_map.get("packageName")),
                    category=normalize_text(row_map.get("category")),
                    description=normalize_text(row_map.get("description")),
                    source_url=normalize_text(row_map.get("sourceUrl")),
                    specifications=parse_specifications(row_map.get("specJson")),
                )
            )
        return components
    finally:
        workbook.close()


def filter_components(
    components: Iterable[ComponentRecord],
    part_number: str | None,
    limit: int | None,
) -> list[ComponentRecord]:
    normalized_part_number = part_number.strip().upper() if part_number else None
    filtered = [
        component
        for component in components
        if normalized_part_number is None or component.part_number == normalized_part_number
    ]
    if limit is not None:
        if limit < 0:
            raise ValueError("--limit must be >= 0")
        filtered = filtered[:limit]
    return filtered


def build_qr_payload(component: ComponentRecord) -> str:
    return f"{{on:,pc:{component.part_number},pm:,qty:,mc:,cc:,pdi:,hp:}}"


def build_secondary_summary(component: ComponentRecord) -> str | None:
    values: list[str] = []
    for key in PREFERRED_SPEC_KEYS:
        value = component.specifications.get(key)
        if value:
            values.append(value.strip())
    for key in sorted(component.specifications):
        if key in PREFERRED_SPEC_KEYS:
            continue
        value = component.specifications[key].strip()
        if value:
            values.append(value)
    deduped: list[str] = []
    seen: set[str] = set()
    for value in values:
        if value and value not in seen:
            deduped.append(value)
            seen.add(value)
    return " · ".join(deduped) or None


def create_qr_bitmap(payload: str, size: int = 720) -> Image.Image:
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=12,
        border=4,
    )
    qr.add_data(payload)
    qr.make(fit=True)
    image = qr.make_image(fill_color="black", back_color="white").convert("RGB")
    return image.resize((size, size), Image.Resampling.NEAREST)


def sanitize_file_stem(value: str) -> str:
    normalized = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    return normalized.strip("._") or "material"


def font_candidates(bold: bool) -> list[Path]:
    if bold:
        return [
            Path("/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc"),
            Path("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"),
            Path("/usr/share/fonts/truetype/noto/NotoSansCJK-Bold.ttc"),
        ]
    return [
        Path("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"),
        Path("/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc"),
        Path("/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf"),
    ]


def load_font(size: int, *, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    for path in font_candidates(bold):
        if path.exists():
            try:
                return ImageFont.truetype(str(path), size=size)
            except OSError:
                continue
    return ImageFont.load_default()


def line_height(font: ImageFont.ImageFont) -> int:
    bbox = font.getbbox("Ag")
    return bbox[3] - bbox[1]


def tokenize_for_wrap(text: str) -> list[str]:
    return re.findall(r"[A-Za-z0-9()./+_%:-]+|\s+|.", text)


def split_token_to_fit(token: str, font: ImageFont.ImageFont, max_width: int) -> list[str]:
    pieces: list[str] = []
    current = ""
    for char in token:
        trial = current + char
        if font.getlength(trial) <= max_width or not current:
            current = trial
            continue
        pieces.append(current)
        current = char
    if current:
        pieces.append(current)
    return pieces or [token]


def wrap_lines(text: str, font: ImageFont.ImageFont, max_width: int, max_lines: int) -> list[str]:
    segments: list[str] = []
    current = ""
    for token in tokenize_for_wrap(text):
        if token.isspace():
            if current and not current.endswith(" "):
                current += " "
            continue
        if font.getlength(token) > max_width:
            split_tokens = split_token_to_fit(token, font, max_width)
        else:
            split_tokens = [token]
        for piece in split_tokens:
            trial = current + piece
            if font.getlength(trial) <= max_width or not current:
                current = trial
                continue
            segments.append(current.rstrip())
            current = piece.lstrip()
            continue
    if current:
        segments.append(current.rstrip())
    if len(segments) <= max_lines:
        return segments
    truncated = segments[:max_lines]
    last = truncated[-1]
    ellipsis = "..."
    while last and font.getlength(last + ellipsis) > max_width:
        last = last[:-1]
    truncated[-1] = (last + ellipsis) if last else ellipsis
    return truncated


def draw_text_block(
    draw: ImageDraw.ImageDraw,
    text: str,
    font: ImageFont.ImageFont,
    fill: str,
    x: float,
    y: float,
    width: int,
    max_lines: int,
) -> float:
    lines = wrap_lines(text, font, width, max_lines)
    if not lines:
        return 0.0
    text_line_height = line_height(font)
    total_height = 0
    for index, line in enumerate(lines):
        draw.text((x, y + total_height), line, font=font, fill=fill)
        total_height += text_line_height
        if index < len(lines) - 1:
            total_height += max(6, math.floor(text_line_height * 0.18))
    return float(total_height)


def create_label_bitmap(component: ComponentRecord, qr_bitmap: Image.Image) -> Image.Image:
    width = 1200
    height = 720
    image = Image.new("RGB", (width, height), BACKGROUND_COLOR)
    draw = ImageDraw.Draw(image)

    card_left = 12
    card_top = 12
    card_right = width - 12
    card_bottom = height - 12
    content_padding = 22
    section_gap = 22
    card_radius = 28
    media_size = height - card_top * 2 - content_padding * 2
    media_left = card_left + content_padding
    media_top = card_top + content_padding
    media_right = media_left + media_size
    media_bottom = media_top + media_size
    qr_inset = 18
    text_left = media_right + section_gap
    text_top = media_top
    text_right = card_right - content_padding
    text_width = text_right - text_left

    draw.rounded_rectangle(
        (card_left, card_top, card_right, card_bottom),
        radius=card_radius,
        fill=CARD_COLOR,
        outline=BORDER_COLOR,
        width=2,
    )
    draw.rounded_rectangle(
        (media_left, media_top, media_right, media_bottom),
        radius=22,
        fill=MEDIA_BG_COLOR,
    )

    qr_render = qr_bitmap.resize(
        (media_size - qr_inset * 2, media_size - qr_inset * 2),
        Image.Resampling.NEAREST,
    )
    image.paste(qr_render, (media_left + qr_inset, media_top + qr_inset))

    title_font = load_font(74, bold=True)
    subtitle_font = load_font(48)
    body_font = load_font(48)

    title = component.name or component.mpn or component.part_number
    subtitle = " · ".join(
        value for value in (component.brand, component.package_name, component.category) if value
    )
    secondary_summary = build_secondary_summary(component)
    detail_line = f"编号: {component.part_number}"

    current_y = float(text_top)
    current_y += draw_text_block(
        draw=draw,
        text=title,
        font=title_font,
        fill=TITLE_COLOR,
        x=float(text_left),
        y=current_y,
        width=text_width,
        max_lines=3,
    )
    if subtitle:
        current_y += 8
        current_y += draw_text_block(
            draw=draw,
            text=subtitle,
            font=subtitle_font,
            fill=SUBTITLE_COLOR,
            x=float(text_left),
            y=current_y,
            width=text_width,
            max_lines=3,
        )
    if secondary_summary:
        current_y += 10
        current_y += draw_text_block(
            draw=draw,
            text=secondary_summary,
            font=body_font,
            fill=BODY_COLOR,
            x=float(text_left),
            y=current_y,
            width=text_width,
            max_lines=3,
        )
    current_y += 12
    draw_text_block(
        draw=draw,
        text=detail_line,
        font=body_font,
        fill=BODY_COLOR,
        x=float(text_left),
        y=current_y,
        width=text_width,
        max_lines=3,
    )
    return image


def output_path_for(component: ComponentRecord, output_dir: Path) -> Path:
    component_id = sanitize_file_stem(component.component_id)
    part_number = sanitize_file_stem(component.part_number)
    return output_dir / f"material_qr_{component_id}_{part_number}.png"


def generate_labels(components: Iterable[ComponentRecord], output_dir: Path) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    generated_paths: list[Path] = []
    for component in components:
        payload = build_qr_payload(component)
        qr_bitmap = create_qr_bitmap(payload)
        label_bitmap = create_label_bitmap(component, qr_bitmap)
        target_path = output_path_for(component, output_dir)
        label_bitmap.save(target_path, format="PNG")
        generated_paths.append(target_path)
    return generated_paths


def main() -> int:
    args = parse_args()
    workbook_path = resolve_input_path(args.input)
    output_dir = resolve_output_path(args.output)
    components = load_components(workbook_path, args.sheet)
    selected = filter_components(
        components=components,
        part_number=args.part_number,
        limit=args.limit,
    )
    if not selected:
        print("No matching components found.")
        return 1
    generated = generate_labels(selected, output_dir)
    print(f"Workbook: {workbook_path}")
    print(f"Sheet: {args.sheet}")
    print(f"Generated: {len(generated)}")
    for path in generated[:10]:
        print(path)
    if len(generated) > 10:
        print(f"... and {len(generated) - 10} more")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
